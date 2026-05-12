# export_results.py: Export training results, checkpoints, and metrics to a local directory.
# Pulls checkpoints, logs, and eval results from the remote server and logs to Excel tracker.

import argparse
import json
import re
import subprocess
import sys
from datetime import datetime
from pathlib import Path

import numpy as np

try:
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill
except ImportError:
    print("openpyxl required: pip install openpyxl")
    sys.exit(1)


_EXPERIMENTS_DIR = Path(__file__).resolve().parent.parent
TRACKER_PATH = _EXPERIMENTS_DIR / "experiment_tracker.xlsx"
RESULTS_DIR = _EXPERIMENTS_DIR / "results"

HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
GOOD_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
WARN_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
BAD_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

COLUMNS = [
    "Run Name", "Task", "Model", "Mode", "Steps", "LR", "Batch",
    "Final Loss", "Eval Episodes", "Overall MAE (rad)",
    "J0 MAE", "J1 MAE", "J2 MAE", "J3 MAE", "J4 MAE", "J5 MAE", "Grip MAE",
    "Dataset", "Num Episodes", "Num Frames",
    "Action Dim", "Cameras", "Backend",
    "Checkpoint Path", "Config", "Server", "Date", "Comments",
]


def parse_train_log(log_path: Path) -> dict:
    """Parse training log for loss curve and final metrics."""
    steps, losses, grad_norms = [], [], []
    with open(log_path) as f:
        for line in f:
            m = re.search(r"Step (\d+): grad_norm=([\d.]+), loss=([\d.]+)", line)
            if m:
                steps.append(int(m.group(1)))
                grad_norms.append(float(m.group(2)))
                losses.append(float(m.group(3)))
    return {
        "steps": steps,
        "losses": losses,
        "grad_norms": grad_norms,
        "final_loss": losses[-1] if losses else None,
        "final_step": steps[-1] if steps else None,
    }


def parse_eval_log(log_path: Path) -> dict:
    """Parse evaluation log for per-episode and overall errors."""
    episodes = []
    overall_mae = None
    per_joint = None

    with open(log_path) as f:
        for line in f:
            m = re.search(r"Ep (\d+): MAE=([\d.]+)\s+joints=\[([\d.,\s]+)\]", line)
            if m:
                ep_idx = int(m.group(1))
                mae = float(m.group(2))
                joints = [float(x.strip()) for x in m.group(3).split(",")]
                episodes.append({"episode": ep_idx, "mae": mae, "joints": joints})

            m2 = re.search(r"Overall MAE: ([\d.]+)", line)
            if m2:
                overall_mae = float(m2.group(1))

    if episodes and overall_mae is None:
        overall_mae = np.mean([e["mae"] for e in episodes])

    if episodes:
        per_joint = np.mean([e["joints"] for e in episodes], axis=0).tolist()

    return {
        "episodes": episodes,
        "overall_mae": overall_mae,
        "per_joint": per_joint,
    }


def fetch_from_server(server: str, remote_path: str, local_path: Path):
    """Rsync a file or directory from the remote server."""
    local_path.parent.mkdir(parents=True, exist_ok=True)
    cmd = ["rsync", "-az", f"{server}:{remote_path}", str(local_path)]
    subprocess.run(cmd, check=True)


def export_run(server: str, run_name: str, config_path: str, fetch_checkpoint: bool = False):
    """Export a single training run's results."""
    run_dir = RESULTS_DIR / run_name
    run_dir.mkdir(parents=True, exist_ok=True)

    # Fetch logs
    print(f"Fetching logs for {run_name}...")
    fetch_from_server(server, "~/openpi/experiments/train/train.log", run_dir / "train.log")
    fetch_from_server(server, "~/eval_out.txt", run_dir / "eval.log")

    # Fetch config
    fetch_from_server(server, f"~/openpi/{config_path}", run_dir / "config.yaml")

    # Optionally fetch checkpoint
    if fetch_checkpoint:
        print(f"Fetching checkpoint (this may take a while)...")
        ckpt_remote = f"~/openpi/checkpoints/pi05_aloha_lipbalm/{run_name}/"
        fetch_from_server(server, ckpt_remote, run_dir / "checkpoint/")

    # Parse logs
    train_metrics = parse_train_log(run_dir / "train.log")
    eval_metrics = parse_eval_log(run_dir / "eval.log")

    # Save parsed metrics as JSON
    metrics = {
        "run_name": run_name,
        "train": train_metrics,
        "eval": eval_metrics,
        "exported_at": datetime.now().isoformat(),
    }
    with open(run_dir / "metrics.json", "w") as f:
        json.dump(metrics, f, indent=2, default=str)

    print(f"Exported to {run_dir}")
    return metrics


def create_loss_plot(run_dir: Path, train_metrics: dict):
    """Create a simple loss curve plot if matplotlib is available."""
    try:
        import matplotlib
        matplotlib.use("Agg")
        import matplotlib.pyplot as plt

        steps = train_metrics["steps"]
        losses = train_metrics["losses"]
        if not steps:
            return None

        fig, (ax1, ax2) = plt.subplots(1, 2, figsize=(14, 5))

        ax1.plot(steps, losses, "b-", linewidth=1.5)
        ax1.set_xlabel("Step")
        ax1.set_ylabel("Loss")
        ax1.set_title("Training Loss")
        ax1.grid(True, alpha=0.3)

        ax2.plot(steps, train_metrics["grad_norms"], "r-", linewidth=1.5)
        ax2.set_xlabel("Step")
        ax2.set_ylabel("Grad Norm")
        ax2.set_title("Gradient Norm")
        ax2.grid(True, alpha=0.3)

        fig.tight_layout()
        plot_path = run_dir / "loss_curve.png"
        fig.savefig(plot_path, dpi=150)
        plt.close(fig)
        return plot_path
    except ImportError:
        return None


def update_tracker(metrics: dict, config: dict, comments: str = ""):
    """Add or update a row in the experiment tracker Excel."""
    if TRACKER_PATH.exists():
        wb = openpyxl.load_workbook(TRACKER_PATH)
    else:
        wb = openpyxl.Workbook()

    # Main results sheet
    if "OpenPI Experiments" not in wb.sheetnames:
        ws = wb.active
        ws.title = "OpenPI Experiments"
        for col_idx, header in enumerate(COLUMNS, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(horizontal="center")
        ws.freeze_panes = "A2"
    else:
        ws = wb["OpenPI Experiments"]

    # Check if run already exists (update in place)
    run_name = metrics["run_name"]
    existing_row = None
    for row in range(2, ws.max_row + 1):
        if ws.cell(row=row, column=1).value == run_name:
            existing_row = row
            break

    row_idx = existing_row or ws.max_row + 1

    train = metrics.get("train", {})
    eval_m = metrics.get("eval", {})
    per_joint = eval_m.get("per_joint", [None] * 7)

    training_cfg = config.get("training", {})
    model_cfg = config.get("model", {})
    data_cfg = config.get("data", {})
    exp_cfg = config.get("experiment", {})
    server_cfg = config.get("server", {})

    values = [
        run_name,
        exp_cfg.get("name", run_name).replace("_pi05_lora", "").replace("_pi05", ""),
        "pi0.5",
        "lora" if training_cfg.get("use_lora", True) else "full",
        train.get("final_step"),
        training_cfg.get("peak_lr", 5e-5),
        training_cfg.get("batch_size"),
        train.get("final_loss"),
        len(eval_m.get("episodes", [])),
        eval_m.get("overall_mae"),
        *[per_joint[i] if i < len(per_joint) else None for i in range(7)],
        data_cfg.get("merged_name"),
        None,  # num_episodes — filled by caller
        None,  # num_frames
        7,  # action_dim
        "cam_high + cam_right_wrist",
        "jax",
        str(RESULTS_DIR / run_name),
        exp_cfg.get("config_file", "config.yaml"),
        server_cfg.get("name", ""),
        datetime.now().strftime("%Y-%m-%d %H:%M"),
        comments,
    ]

    for col_idx, val in enumerate(values, 1):
        cell = ws.cell(row=row_idx, column=col_idx, value=val)
        # Color-code MAE
        if col_idx == 10 and val is not None:  # Overall MAE
            if val < 0.01:
                cell.fill = GOOD_FILL
            elif val < 0.05:
                cell.fill = WARN_FILL
            else:
                cell.fill = BAD_FILL

    # Auto-width columns
    for col_idx in range(1, len(COLUMNS) + 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = max(
            12, len(COLUMNS[col_idx - 1]) + 2
        )

    # Summary sheet
    if "Summary" not in wb.sheetnames:
        wb.create_sheet("Summary")
    summary = wb["Summary"]
    summary["A1"] = "OpenPI Pi0.5 Experiment Tracker"
    summary["A1"].font = Font(bold=True, size=14)
    summary["A2"] = f"Updated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    summary["A3"] = f"Total runs: {ws.max_row - 1}"

    wb.save(TRACKER_PATH)
    print(f"Tracker updated: {TRACKER_PATH}")


def main():
    parser = argparse.ArgumentParser(description="Export OpenPI experiment results")
    parser.add_argument("--server", type=str, default="aws-L40S-server1",
                        help="SSH server name")
    parser.add_argument("--run-name", type=str, default="lipbalm_pi05_lora",
                        help="Training run name")
    parser.add_argument("--config", type=str, default="experiments/configs/lipbalm.yaml",
                        help="Config file path (on server)")
    parser.add_argument("--fetch-checkpoint", action="store_true",
                        help="Also download the model checkpoint")
    parser.add_argument("--comments", type=str, default="",
                        help="Comments for this run")
    args = parser.parse_args()

    # Export results from server
    metrics = export_run(args.server, args.run_name, args.config, args.fetch_checkpoint)

    # Create loss plot
    run_dir = RESULTS_DIR / args.run_name
    train_metrics = metrics.get("train", {})
    plot_path = create_loss_plot(run_dir, train_metrics)
    if plot_path:
        print(f"Loss plot: {plot_path}")

    # Load config for tracker
    import yaml
    config_path = run_dir / "config.yaml"
    if config_path.exists():
        with open(config_path) as f:
            config = yaml.safe_load(f)
    else:
        config = {}

    # Update tracker
    update_tracker(metrics, config, args.comments)

    # Print summary
    eval_m = metrics.get("eval", {})
    print(f"\n{'='*50}")
    print(f"Run: {args.run_name}")
    print(f"Final loss: {train_metrics.get('final_loss')}")
    print(f"Overall MAE: {eval_m.get('overall_mae')}")
    if eval_m.get("per_joint"):
        names = ["j0", "j1", "j2", "j3", "j4", "j5", "grip"]
        for n, v in zip(names, eval_m["per_joint"]):
            print(f"  {n}: {v:.6f}")
    print(f"Results: {run_dir}")
    print(f"Tracker: {TRACKER_PATH}")


if __name__ == "__main__":
    main()
